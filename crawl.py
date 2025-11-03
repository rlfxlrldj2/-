import os, time, random, re
from urllib.parse import urlparse
from datetime import datetime, time as dtime, timedelta

import pytz
import requests
import pandas as pd
from bs4 import BeautifulSoup
from dateutil import parser as dtparser
import trafilatura
from tenacity import retry, wait_exponential, stop_after_attempt
from ratelimit import limits, sleep_and_retry
from tqdm import tqdm

# ===================== 사용자 요건 반영 설정 =====================

# ====== [A] 키워드 세트: 네이버 API 검색용 쿼리 ======
INVEST_KEYWORDS = [
    "투자 협약", "투자협약", "투자 유치", "투자유치", "투자 MOU", "MOU 체결",
    "시설 투자", "시설투자", "설비 투자", "설비투자", "신규 투자", "대규모 투자",
    "공장 투자", "생산기지 투자", "산업단지 투자", "산단 투자",
]
BUILD_KEYWORDS = [
    "신축", "신축공사", "증축", "증축공사", "증개축", "리모델링", "개보수", "대수선",
    "착공", "기공식", "준공", "완공",
    "설계공모", "설계 공모", "입찰 공고", "입찰공고", "시공사 선정", "발주", "공사 발주",
]
INDUSTRIAL_OBJECTS = [
    "공장", "생산시설", "제조시설", "클린룸", "FAB", "팹", "라인", "생산라인",
    "데이터센터", "물류센터", "물류창고",
    "R&D센터", "연구소", "테스트베드", "캠퍼스", "연수원", "본사 사옥", "사옥",
    "산업단지", "산단", "일반산단", "국가산단",
]
INDUSTRIAL_ACTIONS = [
    "설비 증설", "설비증설", "설비 확충", "설비 교체", "설비 도입",
    "라인 증설", "라인 확충", "생산라인 증설",
    "생산능력 확대", "CAPA 확대", "캐파 확대",
    "공장 증설", "공장 신축", "공장 증축", "신공장 건립", "신공장 건설",
]
# (선택) 아래 2개는 필요 시 주석 해제해서 BASE_KEYWORDS에 포함 가능
ARCH_OBJECTS = [
    "건축물", "건물", "빌딩", "주택", "아파트", "오피스텔",
    "학교", "병원", "도서관", "체육관", "공공청사", "청사", "법원",
    "박물관", "미술관", "역사", "터미널", "공항 터미널",
]
LIFECYCLE_TRIGGERS = [
    "개발 계획", "개발계획", "실시설계", "기본설계",
    "도시계획 변경", "용도지역 변경", "산단 지정", "산단 승인",
    "건축허가", "인허가", "사전심의", "환경영향평가",
]

BASE_KEYWORDS = sorted(list(set(
    INVEST_KEYWORDS + BUILD_KEYWORDS + INDUSTRIAL_ACTIONS + INDUSTRIAL_OBJECTS
    # + ARCH_OBJECTS + LIFECYCLE_TRIGGERS
)))
QUERIES = BASE_KEYWORDS[:]   # 네이버 API 검색 쿼리로 사용

PER_QUERY_LIMIT   = 60
ONLY_NAVER_DOMAIN = True
MIN_CHARS         = 200
TIMEZONE          = "Asia/Seoul"

# ===================== 네이버 검색 Open API =====================
API_URL  = "https://openapi.naver.com/v1/search/news.json"
NAVER_CLIENT_ID = os.getenv("NAVER_CLIENT_ID")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET")
if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
    raise SystemExit("Missing NAVER_CLIENT_ID / NAVER_CLIENT_SECRET in env.")

API_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
    "X-Naver-Client-Id": NAVER_CLIENT_ID,
    "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
}
WEB_HEADERS = {k: v for k, v in API_HEADERS.items() if not k.startswith("X-Naver-")}

def compute_week_window_kst():
    """
    지난주 월요일 00:00 ~ 이번주 월요일 00:00 (KST, end-exclusive)
    """
    kst = pytz.timezone(TIMEZONE)
    today_kst = datetime.now(kst).date()
    this_week_mon = today_kst - timedelta(days=today_kst.weekday())
    end_kst = kst.localize(datetime.combine(this_week_mon, dtime(0, 0, 0)))
    start_kst = end_kst - timedelta(days=7)
    return start_kst, end_kst

def _is_naver_host(url: str) -> bool:
    try:
        host = urlparse(url).netloc
        return any(h in host for h in [
            "news.naver.com", "n.news.naver.com", "sports.naver.com", "entertain.naver.com"
        ])
    except Exception:
        return False

@sleep_and_retry
@limits(calls=20, period=60)
@retry(wait=wait_exponential(min=1, max=16), stop=stop_after_attempt(3))
def _search_once(query: str, start: int, display: int, sort: str):
    params = {"query": query, "start": start, "display": display, "sort": sort}
    r = requests.get(API_URL, headers=API_HEADERS, params=params, timeout=20)
    r.raise_for_status()
    return r.json()

def collect_news_items_for_query(query: str, sort: str = "date", limit: int = 200):
    items, start = [], 1
    while len(items) < limit and start <= 1000:
        display = min(100, limit - len(items))
        data = _search_once(query, start=start, display=display, sort=sort)
        batch = data.get("items", [])
        if not batch: break
        items.extend(batch)
        start += len(batch)
        time.sleep(random.uniform(0.2, 0.6))
    return items[:limit]

def collect_all_items(queries, per_query_limit=200):
    all_items = []
    for q in queries:
        try:
            all_items.extend(collect_news_items_for_query(q, sort="date", limit=per_query_limit))
        except Exception as e:
            print("[API 오류]", q, e)
    return all_items

# ===================== 본문 추출 =====================
def _clean_text(s: str) -> str:
    if not s: return ""
    return "\n".join(line.strip() for line in s.splitlines() if line.strip())

def extract_article_text_from_html(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    candidates = ["#dic_area", "#newsct_article", "article"]
    for css in candidates:
        node = soup.select_one(css)
        if not node: continue
        for junk in node.select("script, style, noscript, figure, table, aside, .byline, .copyright"):
            junk.decompose()
        texts = [t.get_text(" ", strip=True) for t in node.find_all(["p","div","span"]) if t.get_text(strip=True)]
        text = _clean_text("\n".join(texts))
        if len(text) > 100: return text
    return _clean_text(soup.get_text(" ", strip=True))

@sleep_and_retry
@limits(calls=30, period=60)
@retry(wait=wait_exponential(min=1, max=16), stop=stop_after_attempt(3))
def fetch_html(url: str):
    r = requests.get(url, headers=WEB_HEADERS, timeout=20, allow_redirects=True)
    r.raise_for_status()
    return r.text, r.url

def extract_fulltext(url: str):
    html, final_url = fetch_html(url)
    text = trafilatura.extract(html, url=final_url, include_comments=False, include_tables=False, favor_recall=True)
    if text and len(text.split()) > 30:
        return _clean_text(text), final_url
    return extract_article_text_from_html(html), final_url

# ===================== 키워드 매칭 =====================
def _make_kw_regex(w: str) -> str:
    parts = []
    for ch in w:
        if ch.isspace(): parts.append(r"\s*")
        elif ch in ("·", "ㆍ"): parts.append(r"[·ㆍ]?")
        else: parts.append(re.escape(ch))
    return "".join(parts)

TAG_PATTERNS = {tag: [re.compile(_make_kw_regex(tag))] for tag in BASE_KEYWORDS}

def find_tags_for_article(title: str, snippet: str, text: str):
    full = f"{title}\n{snippet}\n{text}"
    hit = []
    for tag in BASE_KEYWORDS:
        pats = TAG_PATTERNS[tag]
        if any(p.search(title) for p in pats) or any(p.search(full) for p in pats):
            hit.append(tag)
    return hit

# ====== 2차 필터 & 자동 태깅 ======
EXCLUDE_KEYWORDS = [
    "증권", "주가", "지분", "배당", "리츠", "펀드", "ETF",
    "자산운용", "브로커리지", "리서치센터",
    "대출", "여신", "금리", "채권", "회사채", "유상증자",
    "아파트 분양", "청약", "전세", "월세", "임대료",
    "소프트웨어", "플랫폼", "콘텐츠 제작", "게임 개발",
]
NEGATIVE_SOFT = re.compile("|".join(map(re.escape, EXCLUDE_KEYWORDS)))

ARCH_ACTION_RE = re.compile(r"(신축|증축|증개축|리모델링|개보수|대수선|착공|준공|설계공모|설계\s*공모|입찰\s*공고|시공사\s*선정|발주)")
ARCH_OBJECT_RE = re.compile(r"(건축|건축물|건물|빌딩|주택|아파트|오피스텔|학교|병원|도서관|체육관|공공청사|청사|법원|터미널|역사|공항|물류센터|데이터센터|박물관|미술관|연수원|기숙사|연구동|공장)")

def is_arch_article(title: str, text: str) -> bool:
    full = f"{title}\n{text}"
    if NEGATIVE_SOFT.search(full): return False
    return bool(ARCH_ACTION_RE.search(full) and ARCH_OBJECT_RE.search(full))

INDUSTRIAL_ACTION_RE = re.compile(
    r"(설비\s*(증설|확충|교체|도입)|라인\s*(증설|확충)|생산라인\s*증설|생산능력\s*확대|CAPA\s*확대|캐파\s*확대|공장\s*(신축|증축|증설|건립))"
)
INDUSTRIAL_OBJECT_RE = re.compile(
    r"(공장|생산시설|제조시설|클린룸|FAB|팹|라인|생산라인|데이터센터|물류센터|물류창고|R&D센터|연구소|테스트베드|캠퍼스|연수원|사옥|산업단지|산단)"
)

def is_industrial_facility(title: str, text: str) -> bool:
    full = f"{title}\n{text}"
    if NEGATIVE_SOFT.search(full): return False
    return bool(INDUSTRIAL_ACTION_RE.search(full) and INDUSTRIAL_OBJECT_RE.search(full))

def should_keep_article(title: str, text: str) -> bool:
    return is_arch_article(title, text) or is_industrial_facility(title, text)

TYPE_RULES = {
  "신규 투자":   re.compile(r"(투자\s*(협약|유치)|MOU|시설투자|설비투자|신규\s*투자|대규모\s*투자)"),
  "공장 신축":   re.compile(r"(공장\s*(신축|건립|건설)|신공장)"),
  "시설 증설":   re.compile(r"(증설|확충|라인\s*증설|CAPA\s*확대|캐파\s*확대|생산라인\s*증설|설비\s*(증설|도입|교체))"),
  "입찰/시공":   re.compile(r"(입찰\s*공고|시공사\s*선정|발주)"),
  "준공/완공":   re.compile(r"(준공|완공)"),
}
TARGET_RULES = {
  "공장":       re.compile(r"공장|FAB|팹|클린룸"),
  "데이터센터":  re.compile(r"데이터센터"),
  "물류센터":    re.compile(r"물류센터|물류창고"),
  "연구/R&D":   re.compile(r"R&D센터|연구소|테스트베드"),
  "사옥/본사":   re.compile(r"사옥|본사"),
  "산단":       re.compile(r"산업단지|산단"),
  "기타건축":    re.compile(r"학교|병원|도서관|체육관|청사|역사|터미널|박물관|미술관"),
}

def detect_types(title: str, text: str):
    full = f"{title}\n{text}"
    hits = [k for k, rx in TYPE_RULES.items() if rx.search(full)]
    order = {"신규 투자":0, "공장 신축":1, "시설 증설":2, "입찰/시공":3, "준공/완공":4}
    return sorted(hits, key=lambda x: order.get(x, 99))

def detect_targets(title: str, text: str):
    full = f"{title}\n{text}"
    hits = [k for k, rx in TARGET_RULES.items() if rx.search(full)]
    order = {"공장":0, "데이터센터":1, "물류센터":2, "연구/R&D":3, "사옥/본사":4, "산단":5, "기타건축":6}
    return sorted(hits, key=lambda x: order.get(x, 99))

# ===================== 필드 추출 =====================
PHONE_RE = re.compile(r"(0\d{1,2})[-.\s]?\d{3,4}[-.\s]?\d{4}")

AMOUNT_PATTERNS = [
    re.compile(r"총?\s*(?:투자(?:금액|비)?|사업비|총사업비)\s*[:은]?\s*([\d,\.]+\s*(?:조|억)?\s*원)"),
    re.compile(r"(\d+)\s*조\s*(\d{1,3}(?:,\d{3})*)\s*억\s*원"),
    re.compile(r"([\d,\.]+)\s*조\s*원"),
    re.compile(r"([\d,\.]+)\s*억\s*원"),
]

SCALE_PATTERNS = [
    re.compile(r"규모[^\n]{0,20}?([0-9,\.]+\s*(?:㎡|m²|평|MW|GW|라인|대|톤|t|기|명))"),
    re.compile(r"연면적\s*[:은]?\s*([0-9,\.]+\s*(?:㎡|m²|평))"),
    re.compile(r"(?:건축면적|대지면적|부지면적|부지|대지)\s*[:은]?\s*([0-9,\.]+\s*(?:㎡|m²|평))"),
    re.compile(r"(?:생산|CAPA|캐파)[^\n]{0,10}?([0-9,\.]+\s*(?:톤|t|K|k|만|억|라인))"),
]

LOC_RE = re.compile(
    r"(서울특별시|부산광역시|대구광역시|인천광역시|광주광역시|대전광역시|울산광역시|세종특별자치시|"
    r"경기도|강원특별자치도|강원도|충청북도|충북|충청남도|충남|전라북도|전북|전라남도|전남|경상북도|경북|경상남도|경남|제주특별자치도|제주)"
    r"(?:\s*[가-힣0-9]{1,10}(?:시|군|구))?(?:\s*[가-힣0-9]{1,10}(?:읍|면|동|리))?"
)

def extract_first(patterns, text):
    for p in patterns:
        m = p.search(text)
        if m:
            g = m.groups() if hasattr(m, "groups") else None
            try:
                if g and len(g) > 1: return " ".join([x for x in g if x])
                if g and len(g) == 1: return g[0]
                return m.group(1) if m.lastindex else m.group(0)
            except Exception:
                return m.group(0)
    return ""

def extract_investment(text):
    for p in AMOUNT_PATTERNS:
        m = p.search(text)
        if not m: continue
        if len(m.groups()) >= 2 and m.group(2):
            return f"{m.group(1)}조 {m.group(2)}억 원"
        return m.group(1).strip() if m.groups() else m.group(0).strip()
    return ""

def extract_scale(text): return extract_first(SCALE_PATTERNS, text)
def extract_phone(text):
    m = PHONE_RE.search(text)
    return m.group(0) if m else ""
def extract_location(text):
    m = LOC_RE.search(text)
    return m.group(0) if m else ""
def summarize_100(text):
    if not text: return ""
    s = re.sub(r"\s+", " ", text).strip()
    return s[:100]

# ===================== 메인 파이프라인 =====================
def main_run():
    kst = pytz.timezone(TIMEZONE)
    START_KST, END_KST = compute_week_window_kst()

    items = collect_all_items(QUERIES, PER_QUERY_LIMIT)
    rows = []

    for it in tqdm(items, desc="본문 추출/정리"):
        title = BeautifulSoup(it.get("title", ""), "lxml").get_text(" ", strip=True)
        desc  = BeautifulSoup(it.get("description", ""), "lxml").get_text(" ", strip=True)
        pub   = it.get("pubDate", "")
        try:
            pub_dt = dtparser.parse(pub) if pub else None
            pub_kst = pub_dt.astimezone(kst) if pub_dt and pub_dt.tzinfo else (kst.localize(pub_dt) if pub_dt else None)
        except Exception:
            pub_dt = pub_kst = None

        if pub_kst and not (START_KST <= pub_kst < END_KST):
            continue

        link   = it.get("link") or ""
        origin = it.get("originallink") or ""
        chosen = link if (_is_naver_host(link)) else (link or origin or "")
        if ONLY_NAVER_DOMAIN and not _is_naver_host(chosen):
            if _is_naver_host(link):
                chosen = link
            else:
                continue
        if not chosen:
            continue

        try:
            text, final_url = extract_fulltext(chosen)
        except Exception:
            if origin and origin != chosen:
                try:
                    text, final_url = extract_fulltext(origin)
                except Exception:
                    text, final_url = "", chosen
            else:
                text, final_url = "", chosen

        if len(text) < MIN_CHARS:
            continue

        tags = find_tags_for_article(title, desc, text)
        if not tags:
            continue

        if not should_keep_article(title, text):
            continue

        contact = extract_phone(f"{title}\n{desc}\n{text}")
        types   = detect_types(title, text)
        targets = detect_targets(title, text)

        rows.append({
            "pubDate_kst": pub_kst,
            "title": title,
            "snippet": desc,
            "used_url": final_url,
            "text": text,
            "matched_tags": tags,
            "contact": contact,
            "유형": ", ".join(types),
            "대상시설": ", ".join(targets),
        })
        time.sleep(random.uniform(0.1, 0.3))

    if not rows:
        stamp = datetime.now().strftime("%Y%m%d")
        out_csv = f"naver_news_{stamp}_07to07.csv"
        pd.DataFrame(rows).to_csv(out_csv, index=False, encoding="utf-8-sig")
        out_xlsx = f"naver_news_{stamp}_07to07.report.xlsx"
        with pd.ExcelWriter(out_xlsx) as w:
            pd.DataFrame().to_excel(w, index=False, sheet_name="전체")
        print("[완료] 수집 0건")
        return out_csv, out_xlsx

    df = pd.DataFrame(rows)

    def build_row(row):
        title = str(row.get("title","")).strip()
        text  = str(row.get("text","")).strip()
        snippet = str(row.get("snippet","")).strip()
        full = f"{title}\n{snippet}\n{text}"
        summary_100 = summarize_100(text or snippet)
        scale = extract_scale(full)
        invest = extract_investment(full)
        location = extract_location(full)
        url = row.get("used_url") or ""
        contact = row.get("contact","")
        return pd.Series({
            "헤드라인": title,
            "내용요약": summary_100,
            "규모": scale,
            "투자금액": invest,
            "공사지역": location,
            "담당자 연락처": contact,
            "원본기사 URL 링크": url,
            "유형": row.get("유형",""),
            "대상시설": row.get("대상시설",""),
        })

    norm = df.apply(build_row, axis=1)

    def tags_to_str(tags_list):
        if not isinstance(tags_list, list): return ""
        ordered = [t for t in BASE_KEYWORDS if t in tags_list]
        return ", ".join(ordered)

    df_norm = pd.concat([df[["pubDate_kst","matched_tags"]], norm], axis=1)
    df_norm["키워드"] = df_norm["matched_tags"].apply(tags_to_str)
    df_norm.drop(columns=["matched_tags"], inplace=True)

    df_norm = df_norm.drop_duplicates(subset=["원본기사 URL 링크"]).copy()

    stamp = datetime.now().strftime("%Y%m%d")
    out_csv = f"naver_news_{stamp}_07to07.csv"
    csv_cols = [
        "헤드라인","내용요약","키워드","유형","대상시설",
        "규모","투자금액","공사지역","담당자 연락처","원본기사 URL 링크","pubDate_kst"
    ]
    for c in csv_cols:
        if c not in df_norm.columns: df_norm[c] = ""
    df_norm[csv_cols].to_csv(out_csv, index=False, encoding="utf-8-sig")

    out_xlsx = f"naver_news_{stamp}_07to07_7days.report.xlsx"
    excel_order = [
        "게재시각(KST)","헤드라인","내용요약","키워드","유형","대상시설",
        "규모","투자금액","공사지역","담당자 연락처","원본기사 URL 링크"
    ]

    df_norm_excel = df_norm.copy()
    if "pubDate_kst" in df_norm_excel.columns:
        dt = pd.to_datetime(df_norm_excel["pubDate_kst"], errors="coerce", utc=True)
        dt_kst_naive = dt.dt.tz_convert("Asia/Seoul").dt.tz_localize(None)
        df_norm_excel["게재시각(KST)"] = dt_kst_naive
    else:
        df_norm_excel["게재시각(KST)"] = ""

    for c in excel_order:
        if c not in df_norm_excel.columns: df_norm_excel[c] = ""
    df_norm_excel = df_norm_excel.sort_values(by="게재시각(KST)", ascending=False, kind="mergesort")

    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    header_fill = PatternFill(fill_type="solid", start_color="E2E2E2", end_color="E2E2E2")
    base_font   = Font(name="Malgun Gothic", size=10)
    align_body  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = {
        "A": 18.6, "B": 65.0, "C": 60.0, "D": 20.0, "E": 18.0,
        "F": 20.0, "G": 12.0, "H": 12.0, "I": 22.0, "J": 15.0, "K": 61.5,
    }

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        all_df = df_norm_excel[excel_order].copy()
        all_df.to_excel(writer, index=False, sheet_name="전체")
        ws_all = writer.sheets["전체"]

        for idx, _ in enumerate(excel_order, start=1):
            col_letter = chr(ord('A') + idx - 1)
            ws_all.column_dimensions[col_letter].width = col_widths.get(col_letter, 15.0)

        for cell in ws_all[1]:
            cell.fill = header_fill
            cell.font = base_font
            cell.alignment = align_body
            cell.border = border_all

        for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, min_col=1, max_col=ws_all.max_column):
            for cell in row:
                cell.font = base_font
                cell.alignment = align_body
                cell.border = border_all

        for tag in BASE_KEYWORDS:
            sub = df_norm_excel[df_norm_excel["키워드"].str.contains(
                rf"(?:^|, ){re.escape(tag)}(?:, |$)", na=False)]
            if sub.empty: continue
            sub = sub.sort_values(by="게재시각(KST)", ascending=False, kind="mergesort")
            sub = sub[excel_order]
            name = tag[:31]
            sub.to_excel(writer, index=False, sheet_name=name)

            ws = writer.sheets[name]
            for idx, _ in enumerate(excel_order, start=1):
                col_letter = chr(ord('A') + idx - 1)
                ws.column_dimensions[col_letter].width = col_widths.get(col_letter, 15.0)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = base_font
                cell.alignment = align_body
                cell.border = border_all
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = base_font
                    cell.alignment = align_body
                    cell.border = border_all

    print(f"[완료] 기사 {len(df_norm_excel)}건 → CSV: {out_csv}, XLSX: {out_xlsx}")
    return out_csv, out_xlsx

if __name__ == "__main__":
    try:
        main_run()
    except Exception as e:
        print("[ERROR] Crawler failed:", repr(e))
        import traceback, pandas as pd
        traceback.print_exc()
        from datetime import datetime
        stamp = datetime.now().strftime("%Y%m%d")
        empty_csv = f"naver_news_{stamp}_07to07.csv"
        pd.DataFrame().to_csv(empty_csv, index=False, encoding="utf-8-sig")
        empty_xlsx = f"naver_news_{stamp}_07to07.report.xlsx"
        with pd.ExcelWriter(empty_xlsx) as w:
            pd.DataFrame().to_excel(w, index=False, sheet_name="전체")
        print("[INFO] Wrote empty outputs:", empty_csv, empty_xlsx)
        raise SystemExit(0)
